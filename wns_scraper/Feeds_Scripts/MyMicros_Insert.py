import cx_Oracle
import datetime
import openpyxl
import smtplib
from email.mime.text import MIMEText
from pathlib import Path, PureWindowsPath

try:
    from Feeds_Scripts import MyMicros_Extraction_Config as Conf
    # print(Conf.Cheetah_Last_Logical_Date)
    from Feeds_Scripts import CommonFunction

    Ob_CommonFunction = CommonFunction.SendMail()
except ImportError as e:
    print("Error occured while reading the config file:" + e)
    raise
#below parameters are getting read from configuration file
APPLICATION_NAME=Conf.APPLICATION_NAME
PROCESS_NAME=Conf.PROCESS_NAME
SUBPROCESS_NAME='INSERT_SKS_EUG'
ApplicationPath=Conf.ApplicationPath
ProcessingPath=Conf.ProcessingPath
TargetPath=Conf.TargetPath
LogPath=Conf.LogPath
OraConnectionString=Conf.OraConnectionString
RunningFlag=Conf.RunningFlag
DataLagDays=Conf.DataLagDays
Email_TO_DL=Conf.Email_TO_DL
Alert_TO_DL=Conf.Alert_TO_DL

with open(RunningFlag, 'r') as f:
    LogicalDate = f.readline()
    LogicalDate = datetime.datetime.strptime(LogicalDate, '%Y-%m-%d').date()
print("prepare Insert")

#Current_Date='2018-04-07'
Current_Date=datetime.datetime.today().date()
Current_DTTM=datetime.datetime.today().strftime("%Y-%m-%d@%H.%M.%S")
LogFile=LogPath+'MyMicros_insert_Log'+str(Current_DTTM)+'.txt'
HTMLParseFile=ProcessingPath+'Details_HTMLParseFile_MyMicros_SKS_'+str(LogicalDate)+'.txt'
HTMLParseFile1=ProcessingPath+'Details_HTMLParseFile_MyMicros_EUG_'+str(LogicalDate)+'.txt'
XLWorkbook=TargetPath+'Daily_Extrct_MyMicros_Merge_'+str(LogicalDate)+'.xlsx'
XLWorkSheet_SKS='SKS_Daily_Extrct_'+str(LogicalDate)
XLWorkSheet_EUG='EUG_Daily_Extrct_'+str(LogicalDate)
XLWorkSheet_TotalCount='Total_Count_SKS_EUG_'+str(LogicalDate)

LogFeed=open(LogFile,'w')
LogFeed.write('Connecting Database.. \n')

def SendMail():
    LogFeed.write('Sending final mail.. \n')
    s = smtplib.SMTP(host='smtp-prod.starbucks.net')
    msgString = "MyMicros (SKS and EUG) Extraction has been completed for " + str(LogicalDate) + " at " + str(
        Current_DTTM) + ". Please check today's extracted Excel file\
                    \n\n(" + PureWindowsPath(XLWorkbook).as_uri() + ")\
                    \n\nwith separate worksheets \n\n" + XLWorkSheet_EUG + "\n" + XLWorkSheet_SKS + "\n" + XLWorkSheet_TotalCount + "\
                    \n\nData has been loaded into DEPT_INTL.T_MYMICROSDAILYFEED_HIST and DEPT_INTL.T_MYMICROSDAILYFEEDTOTALCOUNT tables.\
                    \n\nThis is an automated mail ! \n\n\nPlease contact WNS Starbucks EMEA Team for any queries.\n"
    msg = MIMEText(msgString)
    msg['Subject'] = "MyMicros Vitality (SKS and EUG) Extraction has been completed for " + str(LogicalDate) + "."
    msg['From'] = 'prod-wns-EMEA@starbucks.com'
    msg['To'] = Email_TO_DL

    s.sendmail('prod-wns-EMEA@starbucks.com', list(Email_TO_DL.split(',')), msg.as_string())
    s.quit()

def ValidateCycyleDate(query,Type):
    try:
        con = cx_Oracle.connect(OraConnectionString)
        LogFeed.write('Database connection established successfully.. \n')
        cur = con.cursor()
        print(query)

        if (Type=='SELECT'):
            cur.execute(query)
            DateVal = cur.fetchone()
            print(str(DateVal))
            return str(DateVal[0])
        else:
            cur.execute(query)
            cur.execute("COMMIT")
    except cx_Oracle.DatabaseError as e:
        print('Error:' + str(e))
        LogFeed.write("\n Error:' + str(e)\n")
        msgString = "Exiting form the script due to connection issue"
        Subject = "MyMicros process MyMicros_Insert.py DB failed"
        Ob_CommonFunction.SendMail(Alert_TO_DL, msgString, Subject)
        raise
    finally:
        con.close()

def main():
    try:
        con = cx_Oracle.connect(OraConnectionString)
        LogFeed.write('Database connection established successfully.. \n')
        cur = con.cursor()
    except cx_Oracle.DatabaseError as e:
        LogFeed.write('Error:'+str(e))
        raise
    #purge temp table DEPT_INTL.T_MYMACRODAILYFEED
    LogFeed.write('truncate table- DEPT_INTL.T_MYMICROSDAILYFEED.. \n')
    cur.execute("truncate table DEPT_INTL.T_MYMICROSDAILYFEED")

    LogFeed.write('Starting Reading workbook.. \n')
    wb = openpyxl.load_workbook(XLWorkbook,data_only=True)
    ws = wb[XLWorkSheet_SKS]
    LogFeed.write('Reading SKS worksheet.. \n')
    j=1
    for i in range(1, ws.max_row+1):
       Location = ws.cell(row=i, column=j).value
       RevenueCenter = ws.cell(row=i, column=j+1).value
       Check1 = ws.cell(row=i, column=j+2).value
       TransactionTime = ws.cell(row=i, column=j+3).value
       Check2 = str(ws.cell(row=i, column=j+4).value)
       Transaction = str(ws.cell(row=i, column=j+5).value)
       DiscountTotal = ws.cell(row=i, column=j+6).value
       CheckTotal = ws.cell(row=i, column=j+7).value
       Company = ws.cell(row=i, column=j + 8).value
       Extraction_date = ws.cell(row=i, column=j + 9).value
       #LogFeed.write('executing insert into table for SKS.. \n')
       cur.execute(
           "insert into DEPT_INTL.T_MYMICROSDAILYFEED (Location,RevenueCenter,Check1,TransactionTime,Check2,Transaction,DiscountTotal,CheckTotal,Company,Extraction_date) values (:1,:2,:3,:4,:5,:6,:7,:8,:9,:10)",
           (Location, RevenueCenter, Check1, TransactionTime, Check2, Transaction, DiscountTotal, CheckTotal,Company,Extraction_date))
    cur.execute("COMMIT")
    print("SKS Insert done !")
    #wb = openpyxl.load_workbook(XLWorkbook,data_only=True)
    LogFeed.write('Reading EUG worksheet.. \n')
    ws = wb[XLWorkSheet_EUG]
    j=1
    for i in range(1, ws.max_row+1):
       Location = ws.cell(row=i, column=j).value
       RevenueCenter = ws.cell(row=i, column=j+1).value
       Check1 = ws.cell(row=i, column=j+2).value
       TransactionTime = ws.cell(row=i, column=j+3).value
       Check2 = str(ws.cell(row=i, column=j+4).value)
       Transaction = str(ws.cell(row=i, column=j+5).value)
       DiscountTotal = ws.cell(row=i, column=j+6).value
       CheckTotal = ws.cell(row=i, column=j+7).value
       Company = ws.cell(row=i, column=j + 8).value
       Extraction_date = ws.cell(row=i, column=j + 9).value
       #LogFeed.write('executing insert into table for EUG.. \n')
       cur.execute(
           "insert into DEPT_INTL.T_MYMICROSDAILYFEED (Location,RevenueCenter,Check1,TransactionTime,Check2,Transaction,DiscountTotal,CheckTotal,Company,Extraction_date) values (:1,:2,:3,:4,:5,:6,:7,:8,:9,:10)",
           (Location, RevenueCenter, Check1, TransactionTime, Check2, Transaction, DiscountTotal, CheckTotal,Company,Extraction_date))
    #cur.execute("COMMIT")
    print("EUG Insert done !")
    LogFeed.write('Reading Total count worksheet.. \n')
    ws = wb[XLWorkSheet_TotalCount]
    j=1
    for i in range(1, ws.max_row+1):
        Application = ws.cell(row=i, column=j).value
        Total = ws.cell(row=i, column=j+1).value
        Count = ws.cell(row=i, column=j+2).value
        Company = ws.cell(row=i, column=j+3).value
        Extraction_date = ws.cell(row=i, column=j+4).value
        #LogFeed.write('executing insert into table Total Count.. \n')
        cur.execute(
           "insert into DEPT_INTL.T_MYMICROSDAILYFEEDTOTALCOUNT (Application,Total,Count,Company,Extraction_date) values (:1,:2,:3,:4,:5)",
           (Application,Total,Count,Company,Extraction_date))
    #cur.execute("COMMIT")
    print("Total Count Insert done !")
    LogFeed.write('executing History insert.. \n')
    HistInsert='''insert into DEPT_INTL.T_MYMICROSDAILYFEED_HIST 
    select
    trim(Location),trim(RevenueCenter),cast(Check1 as number),to_date(TransactionTime,'MM/DD/YY HH:MI AM') as TransactionDate,
    to_char(to_date(TransactionTime,'MM/DD/YY HH:MI AM'),'HH24:MI:SS') as TrasactionTime,
    trim(Check2),trim(Transaction),DiscountTotal,CheckTotal,Company,to_date(Extraction_date,'YYYY-MM-DD') as Extraction_date,
    current_timestamp(0) as MOD_DT_TM
    from  DEPT_INTL.T_MYMICROSDAILYFEED'''
    cur.execute(HistInsert)
    cur.execute("COMMIT")
    con.commit()

    print("History Insert done !")
    con.close()

    Query = "INSERT INTO DEPT_INTL.TCYCLE_STATUS VALUES('{}','{}','{}',to_date('{}','YYYY-MM-DD'),'COMPLETE',Current_timestamp(0))".format(
        APPLICATION_NAME, PROCESS_NAME, SUBPROCESS_NAME, LogicalDate)
    ValidateCycyleDate(Query, 'INSERT')

    SendMail()

    LogFeed.write('End of the daily insert process.. \n')
    LogFeed.close()

main()