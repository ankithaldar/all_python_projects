import cx_Oracle
import datetime
import openpyxl
import smtplib
from email.mime.text import MIMEText
from pathlib import Path, PureWindowsPath

try:
    from Feeds_Scripts import CHATMyMicros_Conf as Conf
    from Feeds_Scripts import CommonFunction
    Ob_CommonFunction = CommonFunction.SendMail()
    # print(Conf.Cheetah_Last_Logical_Date)
except ImportError as e:
    print("Error occurred while reading the config file:" + e)
    raise
# below parameters are getting read from configuration file
APPLICATION_NAME = Conf.APPLICATION_NAME
PROCESS_NAME = Conf.PROCESS_NAME
# SUBPROCESS_NAME = 'INSERT_CHAT'
ApplicationPath = Conf.ApplicationPath
ProcessingPath = Conf.ProcessingPath
TargetPath = Conf.TargetPath
LogPath = Conf.LogPath
OraConnectionString = Conf.OraConnectionString
RunningFlag = Conf.RunningFlag
DataLagDays = Conf.DataLagDays
Email_TO_DL = Conf.Email_TO_DL
Alert_TO_DL = Conf.Alert_TO_DL
with open(RunningFlag, 'r') as f:
    LogicalDate = f.readline()
    LogicalDate = datetime.datetime.strptime(LogicalDate, '%Y-%m-%d').date()
print("prepare Insert")

# Current_Date = '2018-04-07'
Current_Date = datetime.datetime.today().date()
Current_DTTM = datetime.datetime.today().strftime("%Y-%m-%d@%H.%M.%S")
LogFile = LogPath + 'MyMicros_insert_Log' + str(Current_DTTM) + '.txt'
XLWorkbook = TargetPath + 'Daily_Extrct_MyMicros_CHAT_'+str(LogicalDate) + '.xlsx'
XLWorkSheet_CHAT = 'CHAT_Daily_Extrct_'+str(LogicalDate)
# XLWorkSheet_EUG = 'EUG_Daily_Extrct_'+str(LogicalDate)
XLWorkSheet_TotalCount='Total_Count_CHAT_'+str(LogicalDate)
XLWorkbook1=TargetPath+'Daily_Extrct_MyMicros_CHAT_AVG_' + str(LogicalDate) + '.xlsx'
XLWorkSheet_CHAT1='CHAT_Daily_Extrct_AVG'+str(LogicalDate)
#XLWorkSheet_EUG='EUG_Daily_Extrct_'+str(LogicalDate)
XLWorkSheet_TotalCount1='Total_Count_CHAT_AVG'+str(LogicalDate)


LogFeed=open(LogFile,'w')
LogFeed.write('Connecting Database.. \n')


def ValidateCycyleDate(query,Type):
    try:
        con = cx_Oracle.connect(OraConnectionString)
        LogFeed.write('Database connection established successfully.. \n')
        cur = con.cursor()
        print(query)

        if (Type=='SELECT'):
            cur.execute(query)
            DateVal = cur.fetchone()
            print(str(DateVal))
            return str(DateVal[0])
        else:
            cur.execute(query)
            cur.execute("COMMIT")
    except cx_Oracle.DatabaseError as e:
        print('Error:' + str(e))
        LogFeed.write("\n Error:' + str(e)\n")
        msgString = "\n Error:' + str(e)\n"
        Subject = "MyMicros process CHATMyMicros_INsert.py DB failed"
        Ob_CommonFunction.SendMail(Alert_TO_DL, msgString, Subject)
        raise
    finally:
        con.close()

def main():
    try:
        con = cx_Oracle.connect(OraConnectionString)
        LogFeed.write('Database connection established successfully.. \n')
        cur = con.cursor()
    except cx_Oracle.DatabaseError as e:
        LogFeed.write('Error:'+str(e))
        raise
    # purge temp table DEPT_INTL.T_MYMACRODAILYFEED
    LogFeed.write('truncate table- T_CHATMYMICROSDAILYFEED.. \n')
    cur.execute("truncate table DEPT_INTL.T_CHATMYMICROSDAILYFEED")
    LogFeed.write('truncate table- T_CHATMYMICROSDAILYFEEDAVG.. \n')
    cur.execute("truncate table DEPT_INTL.T_CHATMYMICROSDAILYFEEDAVG")

    LogFeed.write('Starting Reading workbook.. \n')
    wb = openpyxl.load_workbook(XLWorkbook,data_only=True)
    ws = wb[XLWorkSheet_CHAT]
    LogFeed.write('Reading CHAT worksheet.. \n')
    j=1
    for i in range(1, ws.max_row+1):
       Card_Program = ws.cell(row=i, column=j).value
       Gender = ws.cell(row=i, column=j+1).value
       Language = ws.cell(row=i, column=j+2).value
       Not_Activated_Cnt = ws.cell(row=i, column=j+3).value
       Deleted_Cnt = ws.cell(row=i, column=j + 4).value
       Active_Cnt = ws.cell(row=i, column=j + 5).value
       Welcome_Status_Cnt = ws.cell(row=i, column=j + 6).value
       Welcome_Status_Stars = ws.cell(row=i, column=j + 7).value
       Green_Status_Cnt = ws.cell(row=i, column=j + 8).value
       Green_Status_Stars = ws.cell(row=i, column=j + 9).value
       Gold_Status_Cnt = ws.cell(row=i, column=j + 10).value
       Gold_Status_Stars = ws.cell(row=i, column=j + 11).value
       Email_Special_Offers_Cnt = ws.cell(row=i, column=j + 12).value
       Email_Newsletter_Cnt = ws.cell(row=i, column=j + 13).value
       SMS_Special_Offers_Cnt = ws.cell(row=i, column=j + 14).value
       mymicros_file_date = ws.cell(row=i, column=j + 15).value
       mod_date = ws.cell(row=i, column=j + 16).value
       
    
       
       cur.execute(
           "insert into DEPT_INTL.T_CHATMYMICROSDAILYFEED (Card_Program,Gender,Language,Not_Activated_Cnt,Deleted_Cnt,Active_Cnt,Welcome_Status_Cnt,Welcome_Status_Stars,Green_Status_Cnt,Green_Status_Stars,Gold_Status_Cnt,Gold_Status_Stars,Email_Special_Offers_Cnt,Email_Newsletter_Cnt,SMS_Special_Offers_Cnt,mymicros_file_date, mod_date) values (:1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13,:14,:15,:16,:17)", 
           (Card_Program,Gender,Language,Not_Activated_Cnt,Deleted_Cnt,Active_Cnt,Welcome_Status_Cnt,Welcome_Status_Stars,Green_Status_Cnt,Green_Status_Stars,Gold_Status_Cnt,Gold_Status_Stars,Email_Special_Offers_Cnt,Email_Newsletter_Cnt,SMS_Special_Offers_Cnt,mymicros_file_date, mod_date))
    cur.execute("COMMIT")
    print("CHAT Insert done !")
    
    LogFeed.write('executing History insert.. \n')
   
    
    
    HistInsert='''insert into DEPT_INTL.T_CHATMYMICROSDAILYFEED_HIST (Card_Program,Gender,Language,Not_Activated_Cnt,Deleted_Cnt,Active_Cnt,Welcome_Status_Cnt,Welcome_Status_Stars,Green_Status_Cnt,Green_Status_Stars,Gold_Status_Cnt,Gold_Status_Stars,Email_Special_Offers_Cnt,Email_Newsletter_Cnt,SMS_Special_Offers_Cnt,mymicros_file_date, mod_date)
    select
    trim(Card_Program),
    trim(Gender),
    trim(Language),
    Not_Activated_Cnt,
    Deleted_Cnt,
    Active_Cnt,
    Welcome_Status_Cnt,
    Welcome_Status_Stars,
    Green_Status_Cnt,
    Green_Status_Stars,
    Gold_Status_Cnt,
    Gold_Status_Stars,
    Email_Special_Offers_Cnt,
    Email_Newsletter_Cnt,
    SMS_Special_Offers_Cnt,
    to_date(mymicros_file_date,'YYYY-MM-DD') as mymicros_file_date,
    to_date(mod_date,'YYYY-MM-DD') as mod_date
    from  DEPT_INTL.T_CHATMYMICROSDAILYFEED'''
    
    cur.execute(HistInsert)
    cur.execute("COMMIT")
    print("History Insert done !")
    
    TEMP_INSERT = '''insert into DEPT_INTL.T_TEMPCHATMYMICROSDAILYFEED (Not_Activated_Cnt,Deleted_Cnt,Active_Cnt,Welcome_Status_Cnt,Welcome_Status_Stars,Green_Status_Cnt,Green_Status_Stars,Gold_Status_Cnt,Gold_Status_Stars,Email_Special_Offers_Cnt,Email_Newsletter_Cnt,SMS_Special_Offers_Cnt,mymicros_file_date, mod_date)
    select sum(Not_Activated_Cnt),
    sum(Deleted_Cnt),
    sum(Active_Cnt),
    sum(Welcome_Status_Cnt),
    sum(Welcome_Status_Stars),
    sum(Green_Status_Cnt),
    sum(Green_Status_Stars),
    sum(Gold_Status_Cnt),
    sum(Gold_Status_Stars), 
    sum(Email_Special_Offers_Cnt),
    sum(Email_Newsletter_Cnt),
    sum(SMS_Special_Offers_Cnt),
    to_date(mymicros_file_date,'YYYY-MM-DD') as mymicros_file_date,
    to_date(mod_date,'YYYY-MM-DD') as mod_date
    from  DEPT_INTL.T_CHATMYMICROSDAILYFEED_HIST where MYMICROS_FILE_DATE = mymicros_file_date'''
    #cur.execute(TEMP_INSERT)
    #con.commit()

    #print("TEMP Insert done !")
    
    
    
    
    wb1 = openpyxl.load_workbook(XLWorkbook1,data_only=True)
    ws1 = wb1[XLWorkSheet_CHAT1]
    LogFeed.write('Reading CHAT AVG worksheet.. \n')
    j=1
    for i in range(1, ws1.max_row+1):
       Card_Program = ws1.cell(row=i, column=j).value
       Not_Activated_Cnt = ws1.cell(row=i, column=j+1).value
       Deleted_Cnt = ws1.cell(row=i, column=j + 2).value
       Active_Cnt = ws1.cell(row=i, column=j + 3).value
       Welcome_Status_Cnt = ws1.cell(row=i, column=j + 4).value
       Welcome_Status_Stars = ws1.cell(row=i, column=j + 5).value
       Green_Status_Cnt = ws1.cell(row=i, column=j + 6).value
       Green_Status_Stars = ws1.cell(row=i, column=j + 7).value
       Gold_Status_Cnt = ws1.cell(row=i, column=j + 8).value
       Gold_Status_Stars = ws1.cell(row=i, column=j + 9).value
       Email_Special_Offers_Cnt = ws1.cell(row=i, column=j + 10).value
       Email_Newsletter_Cnt = ws1.cell(row=i, column=j + 11).value
       SMS_Special_Offers_Cnt = ws1.cell(row=i, column=j + 12).value
       mymicros_file_date = ws1.cell(row=i, column=j + 13).value
       mod_date = ws1.cell(row=i, column=j + 14).value
       
       #LogFeed.write('executing insert into table for SKS.. \n')
       
       cur.execute(
           "insert into DEPT_INTL.T_CHATMYMICROSDAILYFEEDAVG (Card_Program,Not_Activated_Cnt,Deleted_Cnt,Active_Cnt,Welcome_Status_Cnt,Welcome_Status_Stars,Green_Status_Cnt,Green_Status_Stars,Gold_Status_Cnt,Gold_Status_Stars,Email_Special_Offers_Cnt,Email_Newsletter_Cnt,SMS_Special_Offers_Cnt,mymicros_file_date, mod_date) values (:1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13,:14,:15)", 
           (Card_Program,Not_Activated_Cnt,Deleted_Cnt,Active_Cnt,Welcome_Status_Cnt,Welcome_Status_Stars,Green_Status_Cnt,Green_Status_Stars,Gold_Status_Cnt,Gold_Status_Stars,Email_Special_Offers_Cnt,Email_Newsletter_Cnt,SMS_Special_Offers_Cnt,mymicros_file_date, mod_date))
    cur.execute("COMMIT")
    print("CHAT AVG Insert done !")
   
    LogFeed.write('executing History insert.. \n')
    
    HistInsert1='''insert into DEPT_INTL.T_MYMICROSDAILYFEEDAVG_HIST (Card_Program,Not_Activated_Cnt,Deleted_Cnt,Active_Cnt,Welcome_Status_Cnt,Welcome_Status_Stars,Green_Status_Cnt,Green_Status_Stars,Gold_Status_Cnt,Gold_Status_Stars,Email_Special_Offers_Cnt,Email_Newsletter_Cnt,SMS_Special_Offers_Cnt,mymicros_file_date, mod_date)
    select
    to_date(Card_Program,'YYYY-MM-DD') as Card_Program,
    Not_Activated_Cnt,
    Deleted_Cnt,
    Active_Cnt,
    Welcome_Status_Cnt,
    Welcome_Status_Stars,
    Green_Status_Cnt,
    Green_Status_Stars,
    Gold_Status_Cnt,
    Gold_Status_Stars,
    Email_Special_Offers_Cnt,
    Email_Newsletter_Cnt,
    SMS_Special_Offers_Cnt,
    to_date(mymicros_file_date,'YYYY-MM-DD') as mymicros_file_date,
    to_date(mod_date,'YYYY-MM-DD') as mod_date
    from  DEPT_INTL.T_CHATMYMICROSDAILYFEEDAVG'''
    
    cur.execute(HistInsert1)
    cur.execute("COMMIT")
    con.commit()

    print("History Insert done !")
    
       
    con.close()

    SUBPROCESS_NAME='INSERT_CHAT'
    
    Query = "INSERT INTO DEPT_INTL.TCYCLE_STATUS_CHAT VALUES('{}','{}','{}',to_date('{}','YYYY-MM-DD'),'COMPLETE',Current_timestamp(0))".format(
        APPLICATION_NAME, PROCESS_NAME, SUBPROCESS_NAME, LogicalDate)
    ValidateCycyleDate(Query, 'INSERT')
    
    SUBPROCESS_NAME = 'INSERT_AVERAGE'
    
    Query = "INSERT INTO DEPT_INTL.TCYCLE_STATUS_CHAT VALUES('{}','{}','{}',to_date('{}','YYYY-MM-DD'),'COMPLETE',Current_timestamp(0))".format(
        APPLICATION_NAME, PROCESS_NAME, SUBPROCESS_NAME, LogicalDate)
    ValidateCycyleDate(Query, 'INSERT')
    

    msgString = "CHAT MyMicros Extraction has been completed for " + str(LogicalDate) + " at " + str(
        Current_DTTM) + ". Please check today's extracted Excel file\
                        \n\n(" + PureWindowsPath(XLWorkbook).as_uri() + ")\
                        \n\nwith separate worksheets \n\n" + XLWorkSheet_CHAT + "\
                        \n\nData has been loaded into DEPT_INTL.T_CHATMYMICROSDAILYFEED_HIST table.\
                        \n\nThis is an automated mail ! \n\n\nPlease contact WNS Starbucks EMEA Team for any queries.\n"
    Subject = "CHAT MyMicros Extraction has been completed for " + str(LogicalDate) + "."
    Ob_CommonFunction.SendMail(Email_TO_DL, msgString, Subject)

    LogFeed.write('End of the daily insert process.. \n')
    LogFeed.close()

main()